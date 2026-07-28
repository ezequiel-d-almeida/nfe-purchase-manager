from openpyxl import Workbook
from openpyxl.styles import Font
from models.purchase import Purchase
from pathlib import Path

class ExcelWriter:

    def write(
        self, 
        purchases: list[Purchase], 
        output_path: Path
    ) -> None:

        workbook = Workbook()

        self._create_purchase_sheet(workbook, purchases)

        self._create_installment_sheet(workbook, purchases)

        self._create_product_sheet(workbook, purchases)

        workbook.save(output_path)

        print(f"\nArquivo salvo com sucesso em:\n{output_path}")

    def _create_purchase_sheet(self, workbook, purchases):

        sheet = workbook.active
        sheet.title = "Compras"

        headers = [
            "Fornecedor",
            "CNPJ",
            "Número NF",
            "Data",
            "Valor Total",
            "Qtd Parcelas"
        ]

        for coluna, header in enumerate(headers, start=1):

            cell = sheet.cell(row=1, column=coluna)

            cell.value = header

            cell.font = Font(bold=True)

        linha = 2

        for purchase in purchases:

            sheet.cell(row=linha, column=1).value = purchase.fornecedor
            sheet.cell(row=linha, column=2).value = purchase.cnpj
            sheet.cell(row=linha, column=3).value = purchase.numero_nf
            sheet.cell(row=linha, column=4).value = purchase.data_emissao
            sheet.cell(row=linha, column=5).value = purchase.valor_total
            sheet.cell(row=linha, column=6).value = len(purchase.parcelas)

            linha += 1

    def _create_installment_sheet(self, workbook, purchases):

        sheet = workbook.create_sheet("Parcelas")

        headers = [
            "Número NF",
            "Fornecedor",
            "Parcela",
            "Vencimento",
            "Valor"
        ]

        for coluna, header in enumerate(headers, start=1):

            cell = sheet.cell(row=1, column=coluna)

            cell.value = header

            cell.font = Font(bold=True)

        linha = 2

        for purchase in purchases:

            for installment in purchase.parcelas:

                sheet.cell(row=linha, column=1).value = purchase.numero_nf
                sheet.cell(row=linha, column=2).value = purchase.fornecedor
                sheet.cell(row=linha, column=3).value = installment.numero
                sheet.cell(row=linha, column=4).value = installment.vencimento
                sheet.cell(row=linha, column=5).value = installment.valor

                linha += 1

    def _create_product_sheet(self, workbook, purchases):

        sheet = workbook.create_sheet("Produtos")

        headers = [
            "Número NF",
            "Fornecedor",
            "Código",
            "Produto",
            "Quantidade",
            "Unidade",
            "Valor Unitário",
            "Valor Total"
        ]

        for coluna, header in enumerate(headers, start=1):

            cell = sheet.cell(row=1, column=coluna)

            cell.value = header

            cell.font = Font(bold=True)

        linha = 2

        for purchase in purchases:

            for product in purchase.produtos:

                sheet.cell(row=linha, column=1).value = purchase.numero_nf
                sheet.cell(row=linha, column=2).value = purchase.fornecedor
                sheet.cell(row=linha, column=3).value = product.codigo
                sheet.cell(row=linha, column=4).value = product.descricao
                sheet.cell(row=linha, column=5).value = product.quantidade
                sheet.cell(row=linha, column=6).value = product.unidade
                sheet.cell(row=linha, column=7).value = product.valor_unitario
                sheet.cell(row=linha, column=8).value = product.valor_total

                linha += 1